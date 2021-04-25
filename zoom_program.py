#Şahzı muhterem Sems tarafından bizzat kodlanmıştır, /comprehension pornosu/

from openpyxl import load_workbook
import time
import webbrowser

print("Excel olarak kaydettiğiniz programın adını yazın.")
name_excel = input()
print("Derse giriş gecikme süresi kaç saniye olsun?(Dersi geç açan hocalar için tavsiye edilir...)")
wait_count = input()

wb = load_workbook(name_excel)
ws = wb.active

today = time.strftime("%A")

days_and_rows = {"Monday":1,"Tuesday":3,"Wednesday":5,"Thursday":7,"Friday":9,"Saturday":11,"Sunday":13}
time_row_num = 9
lesson_row_num = time_row_num + 1

with open("zoom_link.txt", "r", encoding="utf-8") as f:
    it = iter(f.readlines())
    lessons_and_links = {x.rstrip("\n"):next(it).rstrip("\n") for x in it}
try:
    time_row = [ws.cell(time_row_num, i).value.strftime("%H:%M") for i in range(2, ws.max_column)]
    lesson_row = [ws.cell(lesson_row_num, i).value for i in range(2, ws.max_column)]
    while True:
        time_now = time.strftime("%H:%M")
        for index, lesson_time in enumerate(time_row):
            if(time_now <= lesson_time):
                lesson_next = lesson_row[index]
                print(f"Sıradaki ders {lesson_next}, başlangıç saati {lesson_time}")
                while True:
                    time.sleep(10)
                    if(time.strftime("%H:%M") == lesson_time):
                        print(f"Planlanmış gecikme başlatılıyor ->{wait_count} saniye")
                        time.sleep(wait_count)
                        print(f"{lesson_next} dersine giriş yapılıyor...")
                        webbrowser.open_new_tab(lessons_and_links[lesson_next])
                        time.sleep(60)
                        break
                break
except:
    print("Bugün için bir programınız bulunmuyor, çıkmak için 'ENTER'a basın.")
    input()
