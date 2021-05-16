#Şahzı muhterem Sems tarafından bizzat kodlanmıştır, /comprehension pornosu/

from openpyxl import load_workbook
import time
import webbrowser

print("excel olarak kaydettiğiniz programın adını yazın.")
name_excel = input()
print("Derse kasıtlı geç girme süresi kaç saniye olsun?(Dersi geç açan hocalar için tavsiye edilir)")
wait_time_count = input()

wb = load_workbook(name_excel)
ws = wb.active

today = time.strftime("%A")

days_and_rows = {"Monday":1,"Tuesday":3,"Wednesday":5,"Thursday":7,"Friday":9,"Saturday":11,"Sunday":13}
time_row_num = 9
lesson_row_num = time_row_num + 1

with open("ders_link.txt", "r", encoding="utf-8") as f:
    it = iter(f.readlines())
    
lessons_and_links = {x.rstrip("\n"):next(it).rstrip("\n") for x in it}


time_row = []
lesson_row = []
for column_num in range(2, ws.max_column + 1):
    if ws.cell(time_row_num, column_num).value == None:
        if ws.cell(lesson_row_num, column_num).value != None:
            print(f"UYARI: Bugünkü derslerinizden {column_num - 1}. ders için bir saat belirtilmemiş, {column_num - 1}. saatteki {ws.cell(lesson_row_num, column_num).value} dersi göz ardı edilicek")
        continue
    time_row.append(ws.cell(time_row_num, column_num).value)
    lesson_row.append(ws.cell(lesson_row_num, column_num).value)
        
if len(time_row) == 0:
    print("Bugünlük dersiniz yok.")
    input()
    exit()

for index, lesson_time in enumerate(time_row):
    time_now = time.strftime("%H:%M")
    try:
        lesson_time = lesson_time.strftime("%H:%M")
    except AttributeError:
        print(f"HATA: Ders saatini yazmanız gereken yere {type(lesson_time)} türünden bir değer girdiğinizi tespit ettik. Lütfen kontrol edin")
        exit()
    if time_now <= lesson_time:
        lesson_next = lesson_row[index]
        print(lesson_next)
        print(f"Sıradaki ders {lesson_next}, başlangıç saati {lesson_time}")
        while True:
            if time.strftime("%H:%M") == lesson_time:
                try:
                    if(lessons_and_links[lesson_next].lower() == "empty"):
                        print(f"'{lesson_next}' dersini 'empty' olarak işaretlediğinizden dolayı ders pas geçildi.")
                        break
                except KeyError:
                    print(f"HATA: Sıradaki ders '{lesson_next}', ders_link.txt'ye kayıtlı değil")
                    exit()
                print(f"Kasıtlı bekleme süresi başlatılıyor ->{wait_time_count} saniye")
                time.sleep(int(wait_time_count))
                print(f"{lesson_next} dersine giriş yapılıyor...")
                webbrowser.open_new_tab(lessons_and_links[lesson_next])
                time.sleep(60)
                break
            time.sleep(10)
print("Bugünlük dersler bitti.")
